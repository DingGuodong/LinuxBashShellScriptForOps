#!/usr/bin/python
# -*- coding: utf-8 -*-
"""
Created by PyCharm.
File Name:              LinuxBashShellScriptForOps:python-get-data-from-excel-using-openpyxl.py
Version:                0.0.1
Author:                 dgden
Author Email:           dgdenterprise@gmail.com
URL:                    https://github.com/DingGuodong/LinuxBashShellScriptForOps
Download URL:           https://github.com/DingGuodong/LinuxBashShellScriptForOps/tarball/master
Create Date:            2019/11/29
Create Time:            14:04
Description:            get data from Excel using openpyxl
Long Description:       row read performance: pandas > xlrd > openpyxl
References:             
Prerequisites:          []
Development Status:     3 - Alpha, 5 - Production/Stable
Environment:            Console
Intended Audience:      System Administrators, Developers, End Users/Desktop
License:                Freeware, Freely Distributable
Natural Language:       English, Chinese (Simplified)
Operating System:       POSIX :: Linux, Microsoft :: Windows
Programming Language:   Python :: 2.6
Programming Language:   Python :: 2.7
Topic:                  Utilities
 """
import openpyxl
from openpyxl.utils import get_column_letter

dest_filename = u'example.xlsx'
wanted_sheet_name = 'example_sheet_name'

wb = openpyxl.load_workbook(dest_filename)

wanted_sheet = wb.get_sheet_by_name(wanted_sheet_name)
max_row = wanted_sheet.max_row  # get max row number
max_column = wanted_sheet.max_column  # get max column number

begin_coordinate = 'A2'
end_coordinate = get_column_letter(max_column) + str(max_row)

for row_objects in wanted_sheet[begin_coordinate:end_coordinate]:
    print [cell.value for cell in row_objects]
